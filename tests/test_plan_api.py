import os
import tempfile
import time
import unittest

from openpyxl import Workbook, load_workbook

from backend_server import create_app
from core.db import get_conn
from core.time_utils import utc_now_iso


class PlanApiTestCase(unittest.TestCase):
    def setUp(self):
        self.app = create_app()
        self.client = self.app.test_client()

    def _create_plan(self):
        payload = {
            "customer_code": "CUST-6002",
            "ship_date": "2026-03-24",
            "merge_mode": "NO_MERGE",
            "orders": [
                {"order_no": "ORD-001", "model": "54-1801", "qty": 120},
                {"order_no": "ORD-002", "model": "54-82202", "qty": 60},
            ],
        }
        created = self.client.post("/api/plans", json=payload)
        self.assertEqual(created.status_code, 201)
        return created.get_json()["plan"]["id"]

    def _calculate_plan_sync(self, plan_id):
        calc = self.client.post("/api/plans/{0}/calculate".format(plan_id))
        self.assertEqual(calc.status_code, 200)
        self.assertEqual(calc.get_json()["solution_count"], 3)

    def test_create_and_calculate_plan(self):
        plan_id = self._create_plan()
        self._calculate_plan_sync(plan_id)

        detail = self.client.get("/api/plans/{0}".format(plan_id))
        self.assertEqual(detail.status_code, 200)
        body = detail.get_json()
        self.assertEqual(body["plan"]["status"], "PENDING_CONFIRM")
        self.assertEqual(len(body["solutions"]), 3)
        self.assertGreater(len(body["solution_item_boxes"]), 0)
        self.assertGreater(len(body["solution_item_pallets"]), 0)
        self.assertIn("order_line_id", body["solution_item_boxes"][0])
        self.assertIn("rule_snapshot_id", body["solution_item_boxes"][0])

    def test_no_merge_disallow_cross_order_carton_and_pallet_mix(self):
        payload = {
            "customer_code": "CUST-NO-MERGE",
            "ship_date": "2026-03-24",
            "merge_mode": "NO_MERGE",
            "orders": [
                {"order_no": "ORD-A", "model": "54-1801", "qty": 6},
                {"order_no": "ORD-B", "model": "54-1801", "qty": 6},
            ],
        }
        created = self.client.post("/api/plans", json=payload)
        self.assertEqual(created.status_code, 201)
        plan_id = created.get_json()["plan"]["id"]

        calc = self.client.post("/api/plans/{0}/calculate".format(plan_id))
        self.assertEqual(calc.status_code, 200)

        detail = self.client.get("/api/plans/{0}".format(plan_id))
        self.assertEqual(detail.status_code, 200)
        body = detail.get_json()
        solution_id = body["solutions"][0]["id"]

        box_rows = [
            row
            for row in body["solution_item_boxes"]
            if row["solution_id"] == solution_id
        ]
        pallet_rows = [
            row
            for row in body["solution_item_pallets"]
            if row["solution_id"] == solution_id
        ]

        carton_order_map = {}
        for row in box_rows:
            carton_order_map.setdefault(row["carton_id"], set()).add(str(row["order_no"]))
        self.assertTrue(carton_order_map)
        for order_set in carton_order_map.values():
            self.assertEqual(len(order_set), 1)

        pallet_order_map = {}
        for row in pallet_rows:
            pallet_order_map.setdefault(row["pallet_id"], set())
            pallet_order_map[row["pallet_id"]].update(carton_order_map.get(row["carton_id"], set()))
        self.assertTrue(pallet_order_map)
        for order_set in pallet_order_map.values():
            self.assertEqual(len(order_set), 1)

    def test_async_confirm_rollback_and_override_upload(self):
        plan_id = self._create_plan()

        queued = self.client.post("/api/plans/{0}/calculate".format(plan_id), json={"async": True})
        self.assertEqual(queued.status_code, 202)
        self.assertTrue(queued.get_json()["queued"])

        solutions = []
        for _ in range(20):
            detail = self.client.get("/api/plans/{0}".format(plan_id))
            self.assertEqual(detail.status_code, 200)
            body = detail.get_json()
            if body["plan"]["status"] == "PENDING_CONFIRM" and body["solutions"]:
                solutions = body["solutions"]
                break
            time.sleep(0.1)

        self.assertGreater(len(solutions), 0)
        selected_solution_id = solutions[0]["id"]

        confirmed = self.client.post(
            "/api/plans/{0}/confirm".format(plan_id),
            json={"solution_id": selected_solution_id, "actor": "qa"},
        )
        self.assertEqual(confirmed.status_code, 200)
        self.assertEqual(confirmed.get_json()["plan"]["status"], "CONFIRMED")
        self.assertEqual(confirmed.get_json()["plan"]["final_solution_id"], selected_solution_id)

        rolled_back = self.client.post(
            "/api/plans/{0}/rollback".format(plan_id),
            json={"reason": "re-check", "actor": "qa"},
        )
        self.assertEqual(rolled_back.status_code, 200)
        self.assertEqual(rolled_back.get_json()["plan"]["status"], "PENDING_CONFIRM")
        self.assertIsNone(rolled_back.get_json()["plan"]["final_solution_id"])

        uploaded = self.client.post(
            "/api/plans/{0}/override-upload".format(plan_id),
            json={
                "file_name": "override_result.xlsx",
                "file_path": "D:/tmp/override_result.xlsx",
                "note": "manual override",
                "actor": "qa",
            },
        )
        self.assertEqual(uploaded.status_code, 201)
        self.assertEqual(uploaded.get_json()["upload"]["file_name"], "override_result.xlsx")

        with tempfile.TemporaryDirectory(prefix="override_upload_") as temp_dir:
            file_path = os.path.join(temp_dir, "override_web_upload.xlsx")
            wb = Workbook()
            wb.save(file_path)
            wb.close()
            with open(file_path, "rb") as fh:
                uploaded2 = self.client.post(
                    "/api/plans/{0}/override-upload".format(plan_id),
                    data={
                        "file": (fh, "override_web_upload.xlsx"),
                        "actor": "qa",
                        "note": "web upload",
                    },
                    content_type="multipart/form-data",
                )
            self.assertEqual(uploaded2.status_code, 201)
            self.assertEqual(uploaded2.get_json()["upload"]["file_name"], "override_web_upload.xlsx")

        final_detail = self.client.get("/api/plans/{0}".format(plan_id))
        self.assertEqual(final_detail.status_code, 200)
        final_body = final_detail.get_json()
        self.assertGreaterEqual(len(final_body["override_uploads"]), 2)
        actions = [row["action"] for row in final_body["audit_logs"]]
        self.assertIn("PLAN_CONFIRM", actions)
        self.assertIn("PLAN_ROLLBACK", actions)
        self.assertIn("PLAN_OVERRIDE_UPLOAD", actions)

    def test_layout_filters_and_export_endpoint(self):
        plan_id = self._create_plan()
        self._calculate_plan_sync(plan_id)

        detail = self.client.get("/api/plans/{0}".format(plan_id))
        self.assertEqual(detail.status_code, 200)
        detail_body = detail.get_json()
        self.assertGreater(len(detail_body["solutions"]), 0)
        selected_solution_id = detail_body["solutions"][0]["id"]

        layout_all = self.client.get("/api/layout/{0}".format(plan_id))
        self.assertEqual(layout_all.status_code, 200)
        layout_body = layout_all.get_json()
        self.assertEqual(layout_body["plan_id"], plan_id)
        self.assertGreater(layout_body["stats"]["row_count"], 0)
        self.assertEqual(layout_body["stats"]["row_count"], len(layout_body["boxes"]))
        # 回归：详情页方案明细依赖 layout 接口返回订单号，避免前端出现 "-"
        self.assertTrue(str(layout_body["boxes"][0].get("order_no", "")).strip())

        first_box = layout_body["boxes"][0]
        pallet_id = first_box["pallet_id"]
        carton_id = first_box["carton_id"]
        model = first_box["models"][0]

        by_pallet = self.client.get("/api/layout/{0}?pallet_id={1}".format(plan_id, pallet_id))
        self.assertEqual(by_pallet.status_code, 200)
        for row in by_pallet.get_json()["boxes"]:
            self.assertEqual(row["pallet_id"], pallet_id)

        by_carton = self.client.get("/api/layout/{0}?carton_id={1}".format(plan_id, carton_id))
        self.assertEqual(by_carton.status_code, 200)
        for row in by_carton.get_json()["boxes"]:
            self.assertEqual(row["carton_id"], carton_id)

        by_model = self.client.get("/api/layout/{0}?model={1}".format(plan_id, model))
        self.assertEqual(by_model.status_code, 200)
        for row in by_model.get_json()["boxes"]:
            self.assertIn(model, row["models"])

        explicit = self.client.get("/api/layout/{0}?solution_id={1}".format(plan_id, selected_solution_id))
        self.assertEqual(explicit.status_code, 200)
        self.assertEqual(explicit.get_json()["solution_id"], selected_solution_id)

        with tempfile.TemporaryDirectory(prefix="plan_export_") as output_dir:
            exported = self.client.post(
                "/api/plans/{0}/export".format(plan_id),
                json={
                    "solution_id": selected_solution_id,
                    "output_dir": output_dir,
                },
            )
            self.assertEqual(exported.status_code, 200)
            self.assertEqual(
                exported.mimetype,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.assertGreater(len(exported.data), 0)

            files = [name for name in os.listdir(output_dir) if name.lower().endswith(".xlsx")]
            self.assertEqual(len(files), 1)

            file_path = os.path.join(output_dir, files[0])
            wb = load_workbook(file_path)
            ws = wb[wb.sheetnames[0]]
            self.assertEqual(str(ws["N4"].value), "2026-03-24")
            self.assertIn("ORD-001", str(ws["A3"].value or ""))
            wb.close()
            exported.close()

    def test_missing_data_should_include_incomplete_box_rule_model(self):
        # 若规则仅有型号但关键字段为空，仍应判定为缺少数据。
        payload = {
            "customer_code": "CUST-MISS",
            "ship_date": "2999-01-02",
            "merge_mode": "NO_MERGE",
            "orders": [{"order_no": "ORD-MISS-001", "model": "480", "qty": 10}],
        }
        created = self.client.post("/api/plans", json=payload)
        self.assertEqual(created.status_code, 201)
        plan_id = created.get_json()["plan"]["id"]

        now = utc_now_iso()
        with get_conn() as conn:
            snapshot_id = conn.execute(
                """
                INSERT INTO rule_snapshot
                (snapshot_type, source_file, version, record_count, payload_preview, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                ("box", "tests_incomplete_480.xlsx", "box_test_incomplete_480", 1, "[]", now),
            ).lastrowid
            conn.execute(
                """
                INSERT INTO rule_model_inner_box
                (snapshot_id, model_code, inner_box_spec, qty_per_carton, gross_weight_kg, raw_payload, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (snapshot_id, "480", None, None, None, "{}", now),
            )
            conn.execute(
                """
                INSERT INTO rule_snapshot_activation
                (snapshot_type, snapshot_id, effective_from, created_at)
                VALUES (?, ?, ?, ?)
                """,
                ("box", snapshot_id, "2999-01-01T00:00:00+00:00", now),
            )

        missing = self.client.get("/api/plans/{0}/missing-data".format(plan_id))
        self.assertEqual(missing.status_code, 200)
        body = missing.get_json()
        self.assertTrue(body["has_missing_data"])
        self.assertIn("480", body["missing_models"])

    def test_save_missing_data_should_sync_to_rule_page_snapshot(self):
        payload = {
            "customer_code": "CUST-SYNC",
            "ship_date": "2999-02-01",
            "merge_mode": "NO_MERGE",
            "orders": [{"order_no": "ORD-SYNC-001", "model": "SYNC-001", "qty": 10}],
        }
        created = self.client.post("/api/plans", json=payload)
        self.assertEqual(created.status_code, 201)
        plan_id = created.get_json()["plan"]["id"]

        now = utc_now_iso()
        with get_conn() as conn:
            snapshot_id = conn.execute(
                """
                INSERT INTO rule_snapshot
                (snapshot_type, source_file, version, record_count, payload_preview, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                ("box", "tests_sync.xlsx", "box_test_sync", 1, "[]", now),
            ).lastrowid
            conn.execute(
                """
                INSERT INTO rule_model_inner_box
                (snapshot_id, model_code, inner_box_spec, qty_per_carton, gross_weight_kg, raw_payload, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (snapshot_id, "SYNC-001", "OLD-BOX", 10, 1.2, "{}", now),
            )
            conn.execute(
                """
                INSERT INTO rule_snapshot_activation
                (snapshot_type, snapshot_id, effective_from, created_at)
                VALUES (?, ?, ?, ?)
                """,
                ("box", snapshot_id, now, now),
            )

        saved = self.client.post(
            "/api/plans/{0}/missing-data".format(plan_id),
            json={
                "box_rules": [
                    {
                        "model_code": "SYNC-001",
                        "inner_box_spec": "NEW-BOX",
                        "qty_per_carton": 25,
                        "gross_weight_kg": 2.5,
                    },
                    {
                        "model_code": "SYNC-NEW",
                        "inner_box_spec": "BOX-NEW",
                        "qty_per_carton": 40,
                        "gross_weight_kg": 4.0,
                    },
                ]
            },
        )
        self.assertEqual(saved.status_code, 200)
        body = saved.get_json()
        self.assertEqual(body["saved_count"], 2)
        self.assertEqual(body["rule_sync"]["snapshot_id"], snapshot_id)
        self.assertEqual(body["rule_sync"]["synced_count"], 2)

        with get_conn() as conn:
            existed = conn.execute(
                """
                SELECT inner_box_spec, qty_per_carton, gross_weight_kg
                FROM rule_model_inner_box
                WHERE snapshot_id = ? AND model_code = ?
                ORDER BY id ASC
                LIMIT 1
                """,
                (snapshot_id, "SYNC-001"),
            ).fetchone()
            added = conn.execute(
                """
                SELECT inner_box_spec, qty_per_carton, gross_weight_kg
                FROM rule_model_inner_box
                WHERE snapshot_id = ? AND model_code = ?
                ORDER BY id ASC
                LIMIT 1
                """,
                (snapshot_id, "SYNC-NEW"),
            ).fetchone()

        # 旧型号应被覆盖，新型号应新增到规则页对应快照
        self.assertIsNotNone(existed)
        self.assertIsNotNone(added)
        self.assertEqual(existed["inner_box_spec"], "NEW-BOX")
        self.assertEqual(int(existed["qty_per_carton"]), 25)
        self.assertAlmostEqual(float(existed["gross_weight_kg"]), 2.5, places=6)
        self.assertEqual(added["inner_box_spec"], "BOX-NEW")
        self.assertEqual(int(added["qty_per_carton"]), 40)
        self.assertAlmostEqual(float(added["gross_weight_kg"]), 4.0, places=6)

    def test_import_template_upload_auto_create_and_calculate(self):
        with tempfile.TemporaryDirectory(prefix="plan_import_") as temp_dir:
            file_path = os.path.join(temp_dir, "orders_import.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "import"
            ws.append(["unit_info", "ship_date", "merge_mode", "need_pallet"])
            ws.append(["CUST-6002", "2026-03-24", "NO_MERGE", "Y"])
            ws.append([])
            ws.append(["model", "customer_line", "qty", "price", "amount", "category", "order_no"])
            ws.append(["54-1801", "LINE-CUST-01", 120, 5.6, 672, "part", "ORD-001"])
            ws.append(["54-82202", "LINE-CUST-02", 60, 4.2, 252, "part", "ORD-002"])
            wb.save(file_path)
            wb.close()

            with open(file_path, "rb") as fh:
                imported = self.client.post(
                    "/api/plans/import",
                    data={
                        "file": (fh, "orders_import.xlsx"),
                        "actor": "qa",
                    },
                    content_type="multipart/form-data",
                )

        self.assertEqual(imported.status_code, 201)
        body = imported.get_json()
        self.assertIn("plan", body)
        self.assertIn("calculate", body)
        self.assertEqual(body["calculate"]["solution_count"], 3)
        self.assertEqual(body["plan"]["status"], "PENDING_CONFIRM")

        plan_id = body["plan"]["id"]
        detail = self.client.get("/api/plans/{0}".format(plan_id))
        self.assertEqual(detail.status_code, 200)
        detail_body = detail.get_json()
        self.assertEqual(detail_body["plan"]["customer_code"], "CUST-6002")
        self.assertEqual(len(detail_body["solutions"]), 3)


if __name__ == "__main__":
    unittest.main()
