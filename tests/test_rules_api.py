import os
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from backend_server import create_app
from core.db import get_conn


def _find_file_by_keyword(project_root, keyword):
    # 在项目根目录按关键字查找 xlsx 文件
    for item in Path(project_root).glob("*.xlsx"):
        if item.name.startswith("~$"):
            continue
        if keyword in item.name:
            return str(item)
    return None


class RulesApiTestCase(unittest.TestCase):
    def setUp(self):
        # 初始化应用与测试客户端，并清理规则数据避免串库
        self.app = create_app()
        self.client = self.app.test_client()
        self.project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self._reset_rule_tables()

    def _reset_rule_tables(self):
        with get_conn() as conn:
            conn.execute("DELETE FROM rule_snapshot_activation")
            conn.execute("DELETE FROM rule_snapshot_conflict")
            conn.execute("DELETE FROM rule_model_inner_box")
            conn.execute("DELETE FROM rule_inner_outer_pallet")
            conn.execute("DELETE FROM rule_snapshot")

    def test_import_pallet_rules_and_get_snapshot(self):
        # 验证托盘规则导入与快照查询
        rule_file = _find_file_by_keyword(self.project_root, "托盘")
        self.assertIsNotNone(rule_file)

        created = self.client.post("/api/rules/pallet/import", json={"file_path": rule_file})
        self.assertEqual(created.status_code, 201)
        snapshot_id = created.get_json()["snapshot_id"]

        detail = self.client.get("/api/rules/snapshots/{0}".format(snapshot_id))
        self.assertEqual(detail.status_code, 200)
        body = detail.get_json()
        self.assertEqual(body["snapshot"]["snapshot_type"], "pallet")
        self.assertGreaterEqual(len(body["records_preview"]), 1)

    def test_import_box_rules_with_xlsx(self):
        # 验证 box 规则导入
        rule_file = _find_file_by_keyword(self.project_root, "导入模板")
        self.assertIsNotNone(rule_file)

        created = self.client.post("/api/rules/box/import", json={"file_path": rule_file})
        self.assertEqual(created.status_code, 201)
        self.assertGreaterEqual(created.get_json()["record_count"], 1)

    def test_import_box_rules_with_multipart_upload(self):
        # 验证网页 multipart 上传导入
        tmp_dir = tempfile.mkdtemp(prefix="rule_upload_")
        file_path = os.path.join(tmp_dir, "box_rules_upload.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "rules"
        ws.append(["model", "inner_box_spec", "qty_per_carton", "gross_weight_kg"])
        ws.append(["UP-001", "105", 20, 12.5])
        wb.save(file_path)
        wb.close()

        with open(file_path, "rb") as fh:
            created = self.client.post(
                "/api/rules/box/import",
                data={"file": (fh, "box_rules_upload.xlsx")},
                content_type="multipart/form-data",
            )

        self.assertEqual(created.status_code, 201)
        body = created.get_json()
        self.assertEqual(body["snapshot_type"], "box")
        self.assertGreaterEqual(body["record_count"], 1)

    def test_conflict_detect_and_activate_version(self):
        # 构造冲突规则：同一型号对应不同内盒
        tmp_dir = tempfile.mkdtemp(prefix="rule_test_")
        file_path = os.path.join(tmp_dir, "box_rules_conflict.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "rules"
        ws.append(["model", "inner_box_spec", "qty_per_carton", "gross_weight_kg"])
        ws.append(["M-001", "104", 20, 12.5])
        ws.append(["M-001", "105", 20, 12.5])
        wb.save(file_path)
        wb.close()

        created = self.client.post("/api/rules/box/import", json={"file_path": file_path})
        self.assertEqual(created.status_code, 201)
        body = created.get_json()
        self.assertEqual(body["snapshot_type"], "box")
        self.assertGreaterEqual(body["conflict_count"], 1)
        snapshot_id = body["snapshot_id"]

        conflicts = self.client.get("/api/rules/snapshots/{0}/conflicts".format(snapshot_id))
        self.assertEqual(conflicts.status_code, 200)
        self.assertGreaterEqual(len(conflicts.get_json()["conflicts"]), 1)

        activated = self.client.post(
            "/api/rules/snapshots/{0}/activate".format(snapshot_id),
            json={"effective_from": "2026-03-24T00:00:00.123Z"},
        )
        self.assertEqual(activated.status_code, 200)

        active = self.client.get("/api/rules/active?snapshot_type=box&at=2026-03-24T12:00:00+00:00")
        self.assertEqual(active.status_code, 200)
        active_body = active.get_json()
        self.assertIsNotNone(active_body["active_snapshot"])
        self.assertEqual(active_body["active_snapshot"]["id"], snapshot_id)

    def test_update_active_rule_record(self):
        # 验证激活规则可通过 API 直接修改并落库
        tmp_dir = tempfile.mkdtemp(prefix="rule_edit_")
        file_path = os.path.join(tmp_dir, "box_rules_edit.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "rules"
        ws.append(["model", "inner_box_spec", "qty_per_carton", "gross_weight_kg"])
        ws.append(["UP-009", "105", 20, 12.5])
        wb.save(file_path)
        wb.close()

        created = self.client.post("/api/rules/box/import", json={"file_path": file_path})
        self.assertEqual(created.status_code, 201)
        snapshot_id = created.get_json()["snapshot_id"]

        activated = self.client.post(
            "/api/rules/snapshots/{0}/activate".format(snapshot_id),
            json={"effective_from": "2026-03-24T00:00:00.123Z"},
        )
        self.assertEqual(activated.status_code, 200)

        detail = self.client.get("/api/rules/snapshots/{0}".format(snapshot_id))
        self.assertEqual(detail.status_code, 200)
        rows = detail.get_json()["records_preview"]
        self.assertGreaterEqual(len(rows), 1)
        record_id = rows[0]["id"]

        updated = self.client.put(
            "/api/rules/active/record",
            json={
                "snapshot_type": "box",
                "snapshot_id": snapshot_id,
                "record_id": record_id,
                "updates": {
                    "inner_box_spec": "205",
                    "qty_per_carton": "66",
                    "gross_weight_kg": "18.2",
                },
            },
        )
        self.assertEqual(updated.status_code, 200)
        updated_row = updated.get_json()["record"]
        self.assertEqual(str(updated_row["inner_box_spec"]), "205")
        self.assertEqual(int(updated_row["qty_per_carton"]), 66)
        self.assertAlmostEqual(float(updated_row["gross_weight_kg"]), 18.2, places=2)


if __name__ == "__main__":
    unittest.main()
