import json
import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from backend_server import create_app
from core.db import get_conn
from services.exporter import export_plan_excel


class ExporterTestCase(unittest.TestCase):
    def setUp(self):
        # 初始化应用，确保迁移已执行
        self.app = create_app()

    def _seed_export_data(self):
        now = "2026-03-25T00:00:00+00:00"
        with get_conn() as conn:
            cursor = conn.execute(
                """
                INSERT INTO shipment_plan
                (customer_code, ship_date, merge_mode, status, source_payload, final_solution_id, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "CUST-EXP",
                    "2026-03-24",
                    "MERGE",
                    "PENDING_CONFIRM",
                    json.dumps({"seed": True}),
                    None,
                    now,
                    now,
                ),
            )
            plan_id = cursor.lastrowid

            solution_cursor = conn.execute(
                """
                INSERT INTO solution
                (plan_id, name, tag, score_rank, box_count, pallet_count, gross_weight_kg, metrics_payload, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    plan_id,
                    "Balanced",
                    "recommended",
                    1,
                    2,
                    2,
                    80.0,
                    json.dumps({"seed": True}),
                    now,
                ),
            )
            solution_id = solution_cursor.lastrowid

            conn.execute(
                "UPDATE shipment_plan SET final_solution_id = ?, status = ? WHERE id = ?",
                (solution_id, "CONFIRMED", plan_id),
            )

            box_rows = [
                ("CARTON-0001", 1, "105", 1, "upright", "M-001", 10, "ORD-001", 12.5),
                ("CARTON-0001", 1, "105", 1, "upright", "M-002", 5, "ORD-002", 12.5),
                ("CARTON-0002", 2, "105", 0, "upright", "M-003", 8, "ORD-003", 11.2),
            ]
            for row in box_rows:
                conn.execute(
                    """
                    INSERT INTO solution_item_box
                    (
                      plan_id, solution_id, carton_id, carton_seq, inner_box_spec,
                      mixed_flag, pose_mode, model_code, qty, order_line_id, order_no,
                      carton_gross_weight_kg, rule_snapshot_id, rule_version, created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        plan_id,
                        solution_id,
                        row[0],
                        row[1],
                        row[2],
                        row[3],
                        row[4],
                        row[5],
                        row[6],
                        None,
                        row[7],
                        row[8],
                        None,
                        None,
                        now,
                    ),
                )

            pallet_rows = [
                ("PALLET-001", 1, 1, "116*116*103", "108*108*90", "CARTON-0001", "upright", "56*38*29", 12.5, 44.0),
                ("PALLET-002", 2, 1, "116*116*103", "108*108*90", "CARTON-0002", "vertical", "56*38*29", 11.2, 41.0),
            ]
            for row in pallet_rows:
                conn.execute(
                    """
                    INSERT INTO solution_item_pallet
                    (
                      plan_id, solution_id, pallet_id, pallet_seq, row_seq,
                      pallet_spec_cm, usable_spec_cm, customized_flag,
                      carton_id, carton_pose, carton_spec_cm, carton_gross_weight_kg,
                      pallet_total_weight_kg, pallet_upright_count, pallet_vertical_count,
                      rule_snapshot_id, rule_version, created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        plan_id,
                        solution_id,
                        row[0],
                        row[1],
                        row[2],
                        row[3],
                        row[4],
                        0,
                        row[5],
                        row[6],
                        row[7],
                        row[8],
                        row[9],
                        1 if row[6] == "upright" else 0,
                        1 if row[6] == "vertical" else 0,
                        None,
                        None,
                        now,
                    ),
                )

        return plan_id, solution_id

    def test_export_mapping_merge_and_pallet_sections(self):
        # 覆盖 W7-01/02/03：映射、合并单元格、托盘分段
        plan_id, solution_id = self._seed_export_data()

        with tempfile.TemporaryDirectory(prefix="export_test_") as output_dir:
            template_path = os.path.join(output_dir, "blank_template.xlsx")
            blank_wb = Workbook()
            blank_wb.save(template_path)
            blank_wb.close()

            result = export_plan_excel(
                plan_id=plan_id,
                solution_id=solution_id,
                template_path=template_path,
                output_dir=output_dir,
            )

            file_path = result["file_path"]
            self.assertTrue(os.path.exists(file_path))

            wb = load_workbook(file_path)
            ws = wb[wb.sheetnames[0]]

            self.assertIsNotNone(ws["A1"].value)
            self.assertIn("ORD-001", str(ws["A3"].value or ""))
            self.assertIn("CUST-EXP", str(ws["A4"].value or ""))
            self.assertEqual(str(ws["N4"].value), "2026-03-24")

            merged = {str(item) for item in ws.merged_cells.ranges}
            self.assertIn("H11:H12", merged)
            self.assertIn("I11:I12", merged)
            self.assertIn("K11:K12", merged)
            self.assertIn("L11:L12", merged)
            self.assertIn("M11:M12", merged)
            self.assertIn("N11:N12", merged)

            self.assertEqual(ws.cell(row=17, column=1).value, "PALLET-001")
            self.assertIn("PALLET-001", str(ws.cell(row=18, column=1).value or ""))
            self.assertIsNone(ws.cell(row=19, column=1).value)
            self.assertIsNone(ws.cell(row=20, column=1).value)
            self.assertIsNone(ws.cell(row=21, column=1).value)
            self.assertEqual(ws.cell(row=22, column=1).value, "PALLET-002")
            wb.close()


if __name__ == "__main__":
    unittest.main()
