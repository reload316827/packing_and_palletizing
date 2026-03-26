from pathlib import Path

from openpyxl import load_workbook
try:
    # .xls 旧格式依赖 xlrd；若环境未安装，允许服务启动，延迟到实际读取时再提示。
    import xlrd  # type: ignore
except Exception:
    xlrd = None


def _load_xlsx_rules(path):
    # .xlsx：遍历全部 sheet，按首行表头构建规则记录。
    workbook = load_workbook(path, data_only=True)
    all_rules = []
    for sheet_name in workbook.sheetnames:
        # 按 sheet 汇总规则，并记录来源行号
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(v).strip() if v is not None else "" for v in rows[0]]
        for idx, row in enumerate(rows[1:], start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            record = {
                headers[i] or "col_{0}".format(i + 1): row[i]
                for i in range(min(len(headers), len(row)))
            }
            record["source_sheet"] = sheet_name
            record["source_row"] = idx
            all_rules.append(record)
    workbook.close()
    return all_rules


def _load_xls_rules(path):
    if xlrd is None:
        # 仅在真的读取 .xls 时才报依赖错误，避免启动阶段崩溃。
        raise RuntimeError("xlrd is required to parse .xls rule files. Install xlrd or provide 装箱.xlsx.")
    workbook = xlrd.open_workbook(str(path))
    all_rules = []
    for sheet in workbook.sheets():
        if sheet.nrows <= 0:
            continue
        headers = [
            str(sheet.cell_value(0, col)).strip()
            if sheet.cell_value(0, col) is not None
            else ""
            for col in range(sheet.ncols)
        ]
        for row_idx in range(1, sheet.nrows):
            row_values = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
            if not any(str(cell).strip() for cell in row_values):
                continue
            record = {
                headers[col] or "col_{0}".format(col + 1): row_values[col]
                for col in range(len(row_values))
            }
            record["source_sheet"] = sheet.name
            record["source_row"] = row_idx + 1
            all_rules.append(record)
    return all_rules


def load_box_rules(file_path):
    # 读取“型号-内盒”规则（支持 .xlsx / .xls），供规则快照导入与自动同步复用。
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError("rule file not found: {0}".format(path))

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _load_xlsx_rules(path)
    if suffix == ".xls":
        return _load_xls_rules(path)
    raise RuntimeError("unsupported file type: {0}".format(suffix))
