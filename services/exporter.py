from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell

from core.db import PROJECT_ROOT, get_conn
from core.errors import AppError


def _safe_name(text):
    value = str(text or "").strip() or "export"
    invalid = set('\\/:*?"<>|')
    return "".join(ch if ch not in invalid else "_" for ch in value)


def _normalize_merge_mode(value):
    text = str(value or "").strip()
    if text in ("MERGE", "拼箱"):
        return "拼箱"
    if text in ("NO_MERGE", "不拼箱"):
        return "不拼箱"
    return text or "不拼箱"


def _pick_solution_id(conn, plan_id, solution_id):
    if solution_id is not None:
        row = conn.execute(
            "SELECT id FROM solution WHERE id = ? AND plan_id = ?",
            (solution_id, plan_id),
        ).fetchone()
        if not row:
            raise AppError("SOLUTION_NOT_FOUND", "solution not found in plan", 404)
        return int(row["id"])

    plan = conn.execute("SELECT * FROM shipment_plan WHERE id = ?", (plan_id,)).fetchone()
    if not plan:
        raise AppError("PLAN_NOT_FOUND", "plan not found: {0}".format(plan_id), 404)

    if plan["final_solution_id"]:
        return int(plan["final_solution_id"])

    row = conn.execute(
        "SELECT id FROM solution WHERE plan_id = ? ORDER BY score_rank ASC, id ASC LIMIT 1",
        (plan_id,),
    ).fetchone()
    if not row:
        raise AppError("SOLUTION_NOT_FOUND", "solution not found in plan", 404)
    return int(row["id"])


def _write_row(ws, row_idx, values):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        # 模板中可能预置合并单元格；若命中非左上角 merged cell，直接跳过，避免导出异常
        if isinstance(cell, MergedCell):
            continue
        cell.value = value


def export_plan_excel(plan_id, solution_id=None, template_path=None, output_dir=None):
    # 直接导出（不读取导出模板）：A1/A3/A4/N4 + 明细区 + 拼箱合并 + 装托分段汇总
    with get_conn() as conn:
        plan = conn.execute("SELECT * FROM shipment_plan WHERE id = ?", (plan_id,)).fetchone()
        if not plan:
            raise AppError("PLAN_NOT_FOUND", "plan not found: {0}".format(plan_id), 404)

        selected_solution_id = _pick_solution_id(conn, plan_id, solution_id)

        box_rows = conn.execute(
            """
            SELECT *
            FROM solution_item_box
            WHERE plan_id = ? AND solution_id = ?
            ORDER BY carton_seq ASC, id ASC
            """,
            (plan_id, selected_solution_id),
        ).fetchall()

        pallet_rows = conn.execute(
            """
            SELECT *
            FROM solution_item_pallet
            WHERE plan_id = ? AND solution_id = ?
            ORDER BY pallet_seq ASC, row_seq ASC, id ASC
            """,
            (plan_id, selected_solution_id),
        ).fetchall()

    workbook = Workbook()
    ws = workbook.active
    ws.title = "导出结果"

    order_nos = sorted({str(row["order_no"] or "") for row in box_rows if row["order_no"]})
    contract_no = "+".join(order_nos) or "-"

    ws["A1"] = _normalize_merge_mode(plan["merge_mode"])
    ws["A3"] = "合同号: {0}".format(contract_no)
    ws["A4"] = "客户代码: {0}".format(plan["customer_code"])
    ws["N4"] = str(plan["ship_date"])

    current_row = 9
    _write_row(ws, current_row, ["装箱明细", "", "", "", "", "", "", ""])
    current_row += 1
    _write_row(ws, current_row, ["订单号", "型号", "数量", "内盒", "外箱号", "外箱规格(cm)", "毛重(kg)", "备注"])
    current_row += 1

    pallet_by_carton = {}
    for row in pallet_rows:
        carton_id = str(row["carton_id"])
        if carton_id not in pallet_by_carton:
            pallet_by_carton[carton_id] = row

    box_group = {}
    for row in box_rows:
        carton_id = str(row["carton_id"])
        box_group.setdefault(carton_id, []).append(row)

    for carton_id, rows in box_group.items():
        start_row = current_row
        pallet_ref = pallet_by_carton.get(carton_id)
        carton_spec = str((pallet_ref and pallet_ref["carton_spec_cm"]) or "56*38*29")

        for row in rows:
            note = "拼箱" if len(rows) > 1 else "-"
            _write_row(
                ws,
                current_row,
                [
                    str(row["order_no"] or "-"),
                    str(row["model_code"] or "-"),
                    int(row["qty"] or 0),
                    str(row["inner_box_spec"] or "-"),
                    carton_id,
                    carton_spec,
                    float(row["carton_gross_weight_kg"] or 0),
                    note,
                ],
            )
            current_row += 1

        if len(rows) > 1:
            end_row = current_row - 1
            for col in [8, 9, 11, 12, 13, 14]:
                ws.merge_cells(start_row=start_row, end_row=end_row, start_column=col, end_column=col)

    current_row += 1
    _write_row(ws, current_row, ["装托明细", "", "", "", "", "", "", ""])
    current_row += 1
    _write_row(ws, current_row, ["托盘号", "外箱号", "外箱规格(cm)", "放置方向", "型号", "数量", "毛重(kg)", "风险"])
    current_row += 1

    rows_by_pallet = {}
    for row in pallet_rows:
        pallet_id = str(row["pallet_id"])
        rows_by_pallet.setdefault(pallet_id, []).append(row)

    for pallet_id, rows in rows_by_pallet.items():
        pallet_spec = str(rows[0]["pallet_spec_cm"] or "116*116*103")
        pallet_weight = float(rows[0]["pallet_total_weight_kg"] or 0)

        for row in rows:
            carton_id = str(row["carton_id"])
            models = sorted({str(item["model_code"] or "-") for item in box_group.get(carton_id, [])})
            qty_sum = sum(int(item["qty"] or 0) for item in box_group.get(carton_id, []))
            pose_text = "侧放+正放" if str(row["carton_pose"] or "upright") == "vertical" else "正放"
            risk_text = "含侧放" if "侧放" in pose_text else "-"

            _write_row(
                ws,
                current_row,
                [
                    pallet_id,
                    carton_id,
                    str(row["carton_spec_cm"] or "56*38*29"),
                    pose_text,
                    "+".join(models) if models else "-",
                    qty_sum,
                    float(row["carton_gross_weight_kg"] or 0),
                    risk_text,
                ],
            )
            current_row += 1

        _write_row(
            ws,
            current_row,
            [
                "托盘汇总 {0}".format(pallet_id),
                "托盘规格 {0}".format(pallet_spec),
                "外箱数 {0}".format(len(rows)),
                "",
                "",
                "",
                "总重 {0:.1f} kg".format(pallet_weight),
                "",
            ],
        )
        current_row += 1

        # 每个托盘段落后保留 3 行空行
        current_row += 3

    export_dir = Path(output_dir) if output_dir else (PROJECT_ROOT / "output" / "exports")
    export_dir.mkdir(parents=True, exist_ok=True)

    file_name = "{0}_{1}_{2}.xlsx".format(
        _safe_name(plan["customer_code"]),
        _safe_name(contract_no or "PLAN"),
        _safe_name(plan["ship_date"]),
    )
    output_path = export_dir / file_name
    workbook.save(str(output_path))

    return {
        "file_path": str(output_path),
        "file_name": file_name,
        "plan_id": plan_id,
        "solution_id": selected_solution_id,
        "row_count_box": len(box_rows),
        "row_count_pallet": len(pallet_rows),
        "download_name_encoded": quote(file_name),
    }
