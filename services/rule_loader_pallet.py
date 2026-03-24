from pathlib import Path

from openpyxl import load_workbook


def load_pallet_rules(file_path):
    # 读取“内盒-外箱-托盘”规则（首个 sheet）
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError("rule file not found: {0}".format(path))
    if path.suffix.lower() != ".xlsx":
        raise RuntimeError("pallet rule file must be .xlsx")

    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    records = []
    for idx, row in enumerate(rows[1:], start=2):
        # 跳过空行，仅保留有效规则记录
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        record = {
            headers[i] or "col_{0}".format(i + 1): row[i]
            for i in range(min(len(headers), len(row)))
        }
        record["source_row"] = idx
        records.append(record)
    return records
