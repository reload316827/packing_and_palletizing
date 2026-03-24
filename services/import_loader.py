from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def load_import_template(file_path: str | Path) -> dict[str, Any]:
    """
    导入模板解析占位实现。
    后续将补充：
    1. 头部字段定位（单位、日期、合并方式、是否打托）
    2. 明细区字段映射
    3. 完整校验与异常报告
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"导入文件不存在: {path}")
    wb = load_workbook(path, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    return {
        "sheet_name": sheet.title,
        "row_count": len(rows),
        "rows_preview": rows[:20],
    }
